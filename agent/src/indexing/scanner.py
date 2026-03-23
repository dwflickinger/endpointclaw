"""File indexing scanner for EndpointClaw agent.

Scans configured monitored paths for files matching monitored extensions,
extracts content, classifies file types, infers project/customer information,
and auto-tags with detected metadata (dollar amounts, materials, addresses).
"""

from __future__ import annotations

import asyncio
import hashlib
import logging
import os
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Optional

import psutil

if TYPE_CHECKING:
    from ..core.config import AgentConfig
    from ..core.database import Database

logger = logging.getLogger("endpointclaw.indexing.scanner")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CONTENT_MAX_CHARS = 5000

EXTENSION_FILE_TYPE_MAP: dict[str, str] = {
    ".jpg": "photo",
    ".jpeg": "photo",
    ".png": "photo",
    ".gif": "photo",
    ".bmp": "photo",
    ".tiff": "photo",
    ".dwg": "drawing",
    ".dxf": "drawing",
}

# Patterns used for auto-tagging
_DOLLAR_RE = re.compile(r"\$[\d,]+(?:\.\d{2})?")
_ADDRESS_RE = re.compile(
    r"\d{1,6}\s+[A-Za-z0-9\s.,#-]{5,60}"
    r"(?:Ave|Avenue|Blvd|Boulevard|Ct|Court|Dr|Drive|Ln|Lane|Pkwy|Parkway"
    r"|Pl|Place|Rd|Road|St|Street|Way|Cir|Circle)\b",
    re.IGNORECASE,
)
_MATERIAL_KEYWORDS = [
    "tpo", "epdm", "metal", "modified bitumen", "mod bit",
    "pvc", "built-up", "bur", "shingle", "slate", "tile",
    "single-ply", "standing seam", "rhinobond", "mechanically attached",
    "fully adhered", "ballasted", "polyiso", "eps", "xps",
    "insulation", "membrane", "flashing", "coping", "scupper",
    "drain", "gutter", "downspout", "penetration", "curb",
    "parapet", "fascia", "soffit", "substrate", "deck",
    "concrete deck", "metal deck", "wood deck", "plywood",
]
_ROOF_TERMS = [
    "roof", "roofing", "re-roof", "reroof", "tear-off",
    "overlay", "recover", "square", "linear feet", "lf",
    "sf", "sq ft", "penetrations",
]


class FileScanner:
    """Scans local directories, extracts content, and indexes files."""

    def __init__(self, config: AgentConfig, database: Database) -> None:
        self.config = config
        self.db = database

        # Progress tracking
        self._total_files: int = 0
        self._processed_files: int = 0
        self._current_file: Optional[str] = None
        self._scanning: bool = False

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def scan_all(self) -> None:
        """Perform a full scan of all monitored paths.

        Yields to user activity when CPU exceeds *config.max_cpu_percent*.
        Logs progress every 100 files.
        """
        self._scanning = True
        self._processed_files = 0
        self._current_file = None

        # Collect all candidate file paths first
        all_paths: list[Path] = []
        monitored_paths: list[str] = getattr(self.config, "monitored_paths", [])
        monitored_extensions: list[str] = getattr(self.config, "monitored_extensions", [])

        logger.info(
            "Starting full scan — paths=%s, extensions=%s",
            monitored_paths,
            monitored_extensions,
        )

        for base_dir in monitored_paths:
            base = Path(base_dir)
            if not base.exists():
                logger.warning("Monitored path does not exist: %s", base)
                continue
            try:
                for root, _dirs, files in os.walk(base):
                    root_path = Path(root)
                    # Skip hidden / system directories
                    if any(
                        part.startswith(".")
                        or part in ("__pycache__", "node_modules", ".git")
                        for part in root_path.parts
                    ):
                        continue
                    for fname in files:
                        fpath = root_path / fname
                        ext = fpath.suffix.lower()
                        if ext in monitored_extensions:
                            all_paths.append(fpath)
            except PermissionError:
                logger.warning("Permission denied walking %s", base_dir)
            except Exception:
                logger.exception("Error walking directory %s", base_dir)

        self._total_files = len(all_paths)
        logger.info("Full scan: found %d candidate files", self._total_files)

        max_cpu: float = getattr(self.config, "max_cpu_percent", 15.0)

        for idx, fpath in enumerate(all_paths, start=1):
            self._current_file = str(fpath)
            try:
                await self.scan_file(fpath)
            except Exception:
                logger.exception("Failed to scan file %s", fpath)

            self._processed_files = idx

            # Progress logging every 100 files
            if idx % 100 == 0:
                logger.info(
                    "Scan progress: %d / %d (%.1f%%)",
                    idx,
                    self._total_files,
                    idx / self._total_files * 100 if self._total_files else 0,
                )

            # Yield to user activity if CPU is high
            try:
                cpu_pct = psutil.cpu_percent(interval=0)
                if cpu_pct > max_cpu:
                    logger.debug(
                        "CPU at %.1f%% (limit %.1f%%) — yielding",
                        cpu_pct,
                        max_cpu,
                    )
                    await asyncio.sleep(1.0)
            except Exception:
                pass

            # Small yield every file to keep event loop responsive
            await asyncio.sleep(0)

        self._current_file = None
        self._scanning = False
        logger.info(
            "Full scan complete — processed %d files", self._processed_files
        )

    async def scan_file(self, file_path: Path) -> dict:
        """Process a single file and return its index record.

        Steps:
        1. Get file stats (size, modified time)
        2. Compute SHA-256 hash
        3. Skip if hash matches existing record
        4. Extract content based on extension
        5. Classify file type
        6. Infer project name from path
        7. Infer customer name
        8. Auto-tag
        9. Return dict with all fields
        """
        file_path = Path(file_path)
        filename = file_path.name
        ext = file_path.suffix.lower()

        # 1. File stats
        stat = await asyncio.to_thread(file_path.stat)
        file_size = stat.st_size
        modified_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()

        # 2. Compute hash
        content_hash = await self._compute_hash(file_path)

        # 3. Check if unchanged
        try:
            existing = await self.db.get_file_by_path(str(file_path))
            if existing and existing.get("content_hash") == content_hash:
                logger.debug("File unchanged (hash match), skipping: %s", file_path)
                return existing
        except Exception:
            # If DB lookup fails, continue with full processing
            logger.debug("Could not check existing record for %s", file_path)

        # 4. Extract content
        content = await self._extract_content(file_path)

        # 5. Classify
        file_type = self._classify_file(filename, content)

        # 6. Infer project
        inferred_project = self._infer_project(file_path)

        # 7. Infer customer
        inferred_customer = self._infer_customer(filename, content)

        # 8. Auto-tag
        tags = self._extract_tags(filename, content)

        record = {
            "file_path": str(file_path),
            "filename": filename,
            "file_size": file_size,
            "modified_at": modified_at,
            "content_hash": content_hash,
            "file_type": file_type,
            "content_extract": content[:CONTENT_MAX_CHARS] if content else "",
            "inferred_project": inferred_project,
            "inferred_customer": inferred_customer,
            "tags": tags,
            "company_id": getattr(self.config, "company_id", "corvex"),
            "indexed_at": datetime.now(timezone.utc).isoformat(),
        }

        # Persist to local database
        try:
            await self.db.upsert_file(record)
        except Exception:
            logger.exception("Failed to upsert file record for %s", file_path)

        return record

    def get_progress(self) -> dict:
        """Return current scan progress."""
        return {
            "scanning": self._scanning,
            "total": self._total_files,
            "processed": self._processed_files,
            "current_file": self._current_file,
            "percent": (
                round(self._processed_files / self._total_files * 100, 1)
                if self._total_files > 0
                else 0.0
            ),
        }

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    async def _compute_hash(self, file_path: Path) -> str:
        """Compute SHA-256 hash of a file in 64 KB chunks."""

        def _hash_sync() -> str:
            sha = hashlib.sha256()
            try:
                with open(file_path, "rb") as f:
                    while True:
                        chunk = f.read(65536)
                        if not chunk:
                            break
                        sha.update(chunk)
            except (OSError, PermissionError) as exc:
                logger.warning("Cannot hash file %s: %s", file_path, exc)
                return ""
            return sha.hexdigest()

        return await asyncio.to_thread(_hash_sync)

    async def _extract_content(self, file_path: Path) -> str:
        """Dispatch content extraction based on file extension."""
        ext = file_path.suffix.lower()
        try:
            if ext == ".xlsx":
                return await asyncio.to_thread(self._extract_xlsx, file_path)
            elif ext == ".pdf":
                return await asyncio.to_thread(self._extract_pdf, file_path)
            elif ext == ".docx":
                return await asyncio.to_thread(self._extract_docx, file_path)
            elif ext in (".csv", ".txt", ".log", ".md", ".json", ".xml"):
                return await asyncio.to_thread(self._extract_text, file_path)
            elif ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff"):
                # Image — just return filename/metadata, no OCR for now
                return f"[Image file: {file_path.name}]"
            elif ext in (".dwg", ".dxf"):
                # Binary CAD formats — filename only
                return f"[CAD drawing: {file_path.name}]"
            else:
                return ""
        except Exception:
            logger.exception("Content extraction failed for %s", file_path)
            return ""

    # -- Format-specific extractors ------------------------------------

    @staticmethod
    def _extract_xlsx(file_path: Path) -> str:
        """Extract cell values from all sheets of an Excel workbook."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed — cannot extract .xlsx content")
            return ""

        parts: list[str] = []
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) for cell in row if cell is not None
                    )
                    if row_text.strip():
                        parts.append(row_text)
                    if len(" ".join(parts)) >= CONTENT_MAX_CHARS:
                        break
                if len(" ".join(parts)) >= CONTENT_MAX_CHARS:
                    break
            wb.close()
        except Exception as exc:
            logger.warning("Error reading xlsx %s: %s", file_path, exc)
        return "\n".join(parts)[:CONTENT_MAX_CHARS]

    @staticmethod
    def _extract_pdf(file_path: Path) -> str:
        """Extract text from all pages of a PDF using PyMuPDF (fitz)."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("PyMuPDF (fitz) not installed — cannot extract .pdf content")
            return ""

        parts: list[str] = []
        try:
            doc = fitz.open(str(file_path))
            for page in doc:
                text = page.get_text()
                if text.strip():
                    parts.append(text)
                if len("\n".join(parts)) >= CONTENT_MAX_CHARS:
                    break
            doc.close()
        except Exception as exc:
            logger.warning("Error reading pdf %s: %s", file_path, exc)
        return "\n".join(parts)[:CONTENT_MAX_CHARS]

    @staticmethod
    def _extract_docx(file_path: Path) -> str:
        """Extract paragraph text from a Word document."""
        try:
            import docx  # python-docx
        except ImportError:
            logger.warning("python-docx not installed — cannot extract .docx content")
            return ""

        parts: list[str] = []
        try:
            document = docx.Document(str(file_path))
            for para in document.paragraphs:
                if para.text.strip():
                    parts.append(para.text)
                if len("\n".join(parts)) >= CONTENT_MAX_CHARS:
                    break
        except Exception as exc:
            logger.warning("Error reading docx %s: %s", file_path, exc)
        return "\n".join(parts)[:CONTENT_MAX_CHARS]

    @staticmethod
    def _extract_text(file_path: Path) -> str:
        """Read a plain-text file directly."""
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read(CONTENT_MAX_CHARS)
        except Exception as exc:
            logger.warning("Error reading text file %s: %s", file_path, exc)
            return ""

    # -- Classification ------------------------------------------------

    @staticmethod
    def _classify_file(filename: str, content: str) -> str:
        """Classify a file into a semantic type based on name and content.

        Priority order (first match wins):
        - estimate/bid/takeoff in name -> "estimate"
        - proposal in name -> "proposal"
        - invoice in name -> "invoice"
        - .jpg/.png -> "photo"
        - measurement/takeoff extension or name -> "measurement"
        - price/rate/cost in name -> "price_sheet"
        - .dwg/.dxf -> "drawing"
        - permit in name -> "permit"
        - letter/email/correspondence in name -> "correspondence"
        - else -> "other"
        """
        name_lower = filename.lower()
        ext = Path(filename).suffix.lower()

        # Extension-based fast path for images and drawings
        if ext in EXTENSION_FILE_TYPE_MAP:
            mapped = EXTENSION_FILE_TYPE_MAP[ext]
            # But override if name has a stronger signal
            if any(kw in name_lower for kw in ("estimate", "bid", "takeoff")):
                return "estimate"
            return mapped

        # Name-based heuristics
        if any(kw in name_lower for kw in ("estimate", "bid", "takeoff")):
            return "estimate"
        if "proposal" in name_lower:
            return "proposal"
        if "invoice" in name_lower:
            return "invoice"
        if any(kw in name_lower for kw in ("measurement", "takeoff", "eagleview", "iroofing")):
            return "measurement"
        if any(kw in name_lower for kw in ("price", "rate", "cost", "pricing")):
            return "price_sheet"
        if "permit" in name_lower:
            return "permit"
        if any(kw in name_lower for kw in ("letter", "email", "correspondence", "memo")):
            return "correspondence"

        # Content-based fallback (lightweight)
        if content:
            content_lower = content[:2000].lower()
            if any(kw in content_lower for kw in ("estimate", "bid amount", "total bid")):
                return "estimate"
            if "proposal" in content_lower and "scope of work" in content_lower:
                return "proposal"
            if "invoice" in content_lower and ("amount due" in content_lower or "total" in content_lower):
                return "invoice"

        return "other"

    # -- Project inference ---------------------------------------------

    @staticmethod
    def _infer_project(file_path: Path) -> Optional[str]:
        """Infer a project name from the file path.

        Looks for directory names after common root directories such as
        Documents, Desktop, Downloads, Estimates, Projects, Jobs.
        """
        parts = file_path.parts
        # Known root markers — the directory *after* one of these is likely
        # a project or client folder.
        markers = {
            "documents", "desktop", "downloads", "estimates",
            "projects", "jobs", "proposals", "clients",
            "onedrive", "sharepoint",
        }
        for idx, part in enumerate(parts):
            if part.lower() in markers and idx + 1 < len(parts):
                candidate = parts[idx + 1]
                # Skip if the candidate is another generic directory
                if candidate.lower() not in markers and not candidate.startswith("."):
                    return candidate

        return None

    # -- Customer inference --------------------------------------------

    @staticmethod
    def _infer_customer(filename: str, content: str) -> Optional[str]:
        """Attempt to infer a customer name from filename or content.

        Simple heuristic: look for patterns like "CustomerName - ...",
        "CustomerName_estimate", or common label prefixes in content.
        """
        # Try filename patterns — "LastName - Something" or "LastName_something"
        name_lower = filename.lower()
        # Strip extension
        stem = Path(filename).stem

        # Pattern: "CustomerName - Description"
        if " - " in stem:
            candidate = stem.split(" - ")[0].strip()
            if len(candidate) > 2 and not any(
                kw in candidate.lower()
                for kw in ("estimate", "proposal", "invoice", "takeoff", "price")
            ):
                return candidate

        # Pattern: "CustomerName_description"
        if "_" in stem:
            candidate = stem.split("_")[0].strip()
            if len(candidate) > 2 and not any(
                kw in candidate.lower()
                for kw in ("estimate", "proposal", "invoice", "takeoff", "price")
            ):
                return candidate

        # Content-based: look for "Customer:", "Client:", "Attn:", "Project for:"
        if content:
            for label in ("customer:", "client:", "attn:", "attention:", "project for:", "prepared for:"):
                idx = content.lower().find(label)
                if idx != -1:
                    after = content[idx + len(label) : idx + len(label) + 100]
                    # Take the first line / first 60 chars
                    line = after.strip().split("\n")[0].strip()
                    if line and len(line) > 2:
                        return line[:60]

        return None

    # -- Auto-tagging --------------------------------------------------

    @staticmethod
    def _extract_tags(filename: str, content: str) -> list[str]:
        """Extract auto-generated tags from filename and content.

        Looks for:
        - Dollar amounts ($X,XXX or $X,XXX.XX)
        - Street addresses
        - Material keywords (TPO, EPDM, metal, modified bitumen, etc.)
        - Roof-related terms
        """
        tags: list[str] = []
        combined = f"{filename} {content}" if content else filename
        combined_lower = combined.lower()

        # Dollar amounts
        dollar_matches = _DOLLAR_RE.findall(combined)
        if dollar_matches:
            tags.append("has_dollar_amounts")
            # Include the largest amount as a tag for quick reference
            amounts: list[float] = []
            for m in dollar_matches:
                try:
                    val = float(m.replace("$", "").replace(",", ""))
                    amounts.append(val)
                except ValueError:
                    pass
            if amounts:
                largest = max(amounts)
                if largest >= 1_000_000:
                    tags.append("amount_1m_plus")
                elif largest >= 100_000:
                    tags.append("amount_100k_plus")
                elif largest >= 10_000:
                    tags.append("amount_10k_plus")

        # Addresses
        if _ADDRESS_RE.search(combined):
            tags.append("has_address")

        # Material keywords
        for kw in _MATERIAL_KEYWORDS:
            if kw in combined_lower:
                tags.append(f"material:{kw}")

        # Roof-related terms
        for term in _ROOF_TERMS:
            if term in combined_lower:
                tags.append(f"roof:{term}")
                break  # One roof tag is enough to flag it

        # Deduplicate while preserving order
        seen: set[str] = set()
        unique_tags: list[str] = []
        for tag in tags:
            if tag not in seen:
                seen.add(tag)
                unique_tags.append(tag)

        return unique_tags
